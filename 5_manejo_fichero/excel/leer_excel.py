#escribir esto en hyper para instalar la libreria pip install openpyxl
from openpyxl import load_workbook #workbook es un libro de excel


def cargar_excel(carpeta, fichero):
    #cargar el fichero de excel en nuestro archivo
    excel = load_workbook(f'./{carpeta}/{fichero}')
    #seleccionar la hoja que quiero. Hoja activa
    hoja = excel.active #la hoja en la que estoy
    #en caso de necesitar seleccionar un hoja concreta lo haremos así:
    #hoja = excel['Finanzas] siendo Finanzas el nombre de la hoja
    lista_empleados = []

    #extraer la primera fila del excel como cabeceras
    filas = hoja.iter_rows (values_only= True) 
    cabeceras = next(filas) #extrae la primera fila, una tupla con las cabeceras

    #recorrer filas y columnas de la hoja de excel
    #for fila in hoja.iter_rows(min_row=2, values_only=True): #empieza en la fila 2 pq la 1ª es la cabecera
        #print(fila) #cada fila es una tupla
    #    for value in fila:
    #        print(value)

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        empleado_dict = dict(zip(cabeceras, fila)) #con esto hace un diccionario
        lista_empleados.append(empleado_dict)
    return lista_empleados


lista = cargar_excel('data', 'empleados.xlsx')
print(lista)