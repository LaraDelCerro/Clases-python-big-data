import csv
from openpyxl import Workbook, load_workbook
import os

def crear_csv(lista, nombre, carpeta):
    fichero = open(f"./{carpeta}/{nombre}", 'a', encoding='UTF-8', newline='')
    # saber las cabeceras.
    cabeceras = []
    for key in lista[0].keys():
        cabeceras.append(key)
    #volcar los datos de la lista en un objeto csv
    mi_csv = csv.DictWriter(fichero, fieldnames=cabeceras) 
    #imprimir cabeceras en la primera linea
    mi_csv.writeheader()
    # Escribir cada fila en el csv
    mi_csv.writerows(lista)
    # cerramos el fichero
    fichero.close()
        

def crear_excel(lista, carpeta, fichero, hoja  ):
    if os.path.exists(f"./{carpeta}/{fichero}"):
        wb = load_workbook(f"./{carpeta}/{fichero}")
    else:
        #crear el libro de excel
        wb = Workbook()
        wb.remove(wb.active)

    #seleccionamos la primera hoja
    hoja = wb.create_sheet(title=hoja )
    #hoja.title = 'Trabajadores'

    #extraer de un diccionario cq de mi lista las cabeceras
    cabeceras =list(lista[0].keys())
    #añado las cabeceras a mi hoja
    hoja.append(cabeceras)

    #recorrer la lista de datos para imprimir en cada fila un dato concreto
    for item in lista:
        #para que esto funcione el empleado tiene que tener los datos en el mismo orden que la lista de cabeceras y estar convertido en lista
        #lista_empleado = list(empleado.values())
        lista = [item[clave] for clave in cabeceras]#para cada empleado crea una lista con toda su info
        hoja.append(lista)#añadimos a la hoja de excel la lista de cada empleado

    wb.save(f'./{carpeta}/{fichero}')#guardar 