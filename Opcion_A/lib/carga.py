from openpyxl import load_workbook, Workbook 
import csv
import json
import xml.etree.ElementTree as et


def cargar_csv(carpeta, fichero):
    """
This is an example of Google style.

Args:
    param1: This is the first param.
    param2: This is a second param.

Returns:
    This is a description of what is returned.

Raises:
    KeyError: Raises an exception.
"""
    fichero = open(f'{carpeta}/{fichero}', 'r', encoding='UTF-8')
    lector = csv.DictReader(fichero) #la fcón DictReader mete todos los registros en un diccionario, identifica tb las cabeceras
    lista = list(lector) #transformo en lista el cjto de diccionarios que tengo en la vble lector. 
    fichero.close()
    return lista
    





def cargar_excel(carpeta, fichero):
    #cargar el fichero de excel en nuestro archivo
    excel = load_workbook(f'./{carpeta}/{fichero}')
    hoja = excel.active 
    lista = []

    #extraer la primera fila del excel como cabeceras
    filas = hoja.iter_rows (values_only= True) 
    cabeceras = next(filas) #extrae la primera fila, una tupla con las cabeceras


    for fila in hoja.iter_rows(min_row=2, values_only=True):
        diccionario = dict(zip(cabeceras, fila))
        lista.append(diccionario)
    return lista  

   



def cargar_json(carpeta, fichero):
    try:
        fichero = open(f'./{carpeta}/{fichero}', "r", encoding="UTF-8")
        datos = json.load(fichero)
        return datos
    except FileNotFoundError:
        print('Archivo o carpeta no encontrado')





# def cargar_xml(carpeta, fichero):
#     fichero = et.parse(f'./{carpeta}/{fichero}')
#     nodo_raiz = fichero.getroot() #me devuelve el nodo ppal. 
#     lista = []

#     for elemento in nodo_raiz.findall('patrocinador'): #findall para poder recorrerlo como un array
#         #extraer la información del empleado e insertarla en diccionario empleado
#         patrocinador = {}
#         patrocinador = {
#             'nombre_empresa' : elemento.get('nombre_empresa'),
#             'contacto': elemento.find('contacto').text,
#             'email': elemento.find('email').text,
#             'importe_patrocinio': elemento.find('importe_patrocinio').text,
#             'categoria': elemento.find('categoria').text,
#             'fecha_inicio': elemento.find('fecha_inicio').text,
#             'fecha_fin': elemento.find('fecha_fin').text

#         }
#         lista.append(patrocinador)
#     return lista







def cargar_xml(carpeta, fichero):
    fichero = et.parse(f'./{carpeta}/{fichero}')
    nodo_raiz = fichero.getroot() #me devuelve el nodo ppal. 
    lista = []
    
    conjunto = set()
    etiquetas = []
    for elemento in nodo_raiz.findall('patrocinador'): #findall para poder recorrerlo como un array
        patrocinador = {}
        for subitem in elemento:
            conjunto.add(subitem.tag)
            etiquetas = list(conjunto)
        for etiqueta in etiquetas:
            patrocinador[etiqueta] = elemento.find(etiqueta).text
        lista.append(patrocinador)
    
    return lista  

    
    

def pintar_info_fichero(lista, fichero, num_registros):
    cabeceras =lista[0].keys()
    print(f'## INFO DEL ARCHIVO {fichero} ##')
    print('_'*30)
    print(f'El nombre del fichero es: {fichero}')
    print(' ')
    print(f'El número total de registros es {len(lista)}')
    print(' ')
    print(f'Los nombres de los campos son {list(cabeceras)}')
    print(' ')
    contador = 1
    while contador <= num_registros:
        for registro in lista[:num_registros]:
            print(f'Registro {contador}: {registro}')
            print(' ')
            contador +=1

            

                
                  
   
    
         



   