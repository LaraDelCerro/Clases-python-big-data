#Generar un archivo excel inventario_limpio.xlsx
from openpyxl import load_workbook, Workbook 
from lib.funciones import limpiar_id, limpiar_precio, limpiar_stock, limpiar_string



def cargar_excel(carpeta, fichero):
    #cargar el fichero de excel en nuestro archivo
    excel = load_workbook(f'./{carpeta}/{fichero}')
    #seleccionar la hoja que quiero. Hoja activa
    hoja = excel.active #la hoja en la que estoy
    #en caso de necesitar seleccionar un hoja concreta lo haremos así:
    #hoja = excel['Finanzas] siendo Finanzas el nombre de la hoja
    lista_productos = []

    #extraer la primera fila del excel como cabeceras
    filas = hoja.iter_rows (values_only= True) 
    cabeceras = next(filas) #extrae la primera fila, una tupla con las cabeceras

    #recorrer filas y columnas de la hoja de excel
    #for fila in hoja.iter_rows(min_row=2, values_only=True): #empieza en la fila 2 pq la 1ª es la cabecera
        #print(fila) #cada fila es una tupla
    #    for value in fila:
    #        print(value)

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        producto_dict = dict(zip(cabeceras, fila)) #con esto hace un diccionario
        producto_dict['id_producto'] = limpiar_id(producto_dict['id_producto'])
        producto_dict['categoria'] = limpiar_string(producto_dict['categoria'], True)
        producto_dict['nombre'] = limpiar_string(producto_dict['nombre'])
        producto_dict['precio'] = limpiar_precio(producto_dict['precio'])
        producto_dict['stock'] = limpiar_stock(producto_dict['stock'])
        print(producto_dict)
        

        lista_productos.append(producto_dict)
    return lista_productos



def crear_excel(carpeta, fichero, datos):
    #crear el libro de excel
    wb = Workbook()
    #seleccionamos la primera hoja
    hoja = wb.active
    hoja.title = 'Trabajadores'

    #extraer de un diccionario cq de mi lista las cabecerasd
    cabeceras =list( datos[0].keys())
    #añado las cabeceras a mi hoja
    hoja.append(cabeceras)

    #recorrer la lista de datos para imprimir en cada fila un dato concreto
    for producto in datos:
        #para que esto funcione el empleado tiene que tener los datos en el mismo orden que la lista de cabeceras y estar convertido en lista
        #lista_empleado = list(empleado.values())
        lista_producto = [producto[clave] for clave in cabeceras]#para cada empleado crea una lista con toda su info
        #print(lista_empleado)
        hoja.append(lista_producto)#añadimos a la hoja de excel la lista de cada empleado


    wb.save(f'./{carpeta}/{fichero}')#guardar 

 
lista = cargar_excel('data', 'inventario_sucio.xlsx')
print(lista)

inventario_limpio= crear_excel('data','inventario_limpio.xlsx', lista)